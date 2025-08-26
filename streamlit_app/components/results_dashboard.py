#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Results dashboard komponenty pre ICO Collector Streamlit App
"""

import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime
from utils.config import CHART_COLORS

def render_main_results_dashboard(results: Dict[str, List], df: pd.DataFrame):
    """
    Hlavný dashboard s výsledkami spracovania.
    """
    if not results:
        st.warning("⚠️ Žiadne výsledky na zobrazenie")
        return
    
    st.header("📊 Dashboard výsledkov")
    
    # Základné štatistiky
    render_summary_metrics(results)
    
    # Hlavné grafy
    col1, col2 = st.columns(2)
    
    with col1:
        render_success_rate_chart(results)
    
    with col2:
        render_match_strategy_chart(results)
    
    # Pokročilé analýzy
    render_advanced_analytics(results, df)
    
    # Exportné možnosti
    render_export_section(results, df)

def render_summary_metrics(results: Dict[str, List]):
    """
    Vykresli súhrnné metriky.
    """
    total_companies = len(results.get('ICO', []))
    successful_matches = sum(1 for ico in results.get('ICO', []) if ico)
    failed_searches = total_companies - successful_matches
    success_rate = (successful_matches / total_companies) * 100 if total_companies > 0 else 0
    
    st.subheader("📈 Súhrnné štatistiky")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(
            "Celkovo spracované",
            total_companies,
            help="Celkový počet firiem spracovaných z Excel súboru"
        )
    
    with col2:
        st.metric(
            "Úspešné zhody",
            successful_matches,
            delta=f"{success_rate:.1f}%",
            help="Počet firiem, pre ktoré bolo nájdené IČO"
        )
    
    with col3:
        st.metric(
            "Neúspešné vyhľadávania",
            failed_searches,
            delta=f"-{100-success_rate:.1f}%",
            delta_color="inverse",
            help="Počet firiem, pre ktoré nebolo možné nájsť IČO"
        )
    
    with col4:
        # Gauge pre úspešnosť
        gauge_fig = create_success_gauge(success_rate)
        st.plotly_chart(gauge_fig, use_container_width=True)

def create_success_gauge(success_rate: float):
    """
    Vytvorí gauge chart pre úspešnosť.
    """
    fig = go.Figure(go.Indicator(
        mode = "gauge+number",
        value = success_rate,
        domain = {'x': [0, 1], 'y': [0, 1]},
        title = {'text': "Úspešnosť"},
        number = {'suffix': "%"},
        gauge = {
            'axis': {'range': [None, 100]},
            'bar': {'color': CHART_COLORS['primary']},
            'steps': [
                {'range': [0, 50], 'color': CHART_COLORS['error']},
                {'range': [50, 80], 'color': CHART_COLORS['warning']},
                {'range': [80, 100], 'color': CHART_COLORS['success']}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    
    fig.update_layout(height=200, margin=dict(l=20, r=20, t=40, b=20))
    return fig

def render_success_rate_chart(results: Dict[str, List]):
    """
    Vykresli graf úspešnosti.
    """
    total = len(results.get('ICO', []))
    successful = sum(1 for ico in results.get('ICO', []) if ico)
    failed = total - successful
    
    fig = px.pie(
        values=[successful, failed],
        names=['Úspešné', 'Neúspešné'],
        title='Rozdelenie úspešnosti vyhľadávania',
        color_discrete_map={
            'Úspešné': CHART_COLORS['success'],
            'Neúspešné': CHART_COLORS['error']
        },
        hole=0.4
    )
    
    fig.update_traces(
        textposition='inside',
        textinfo='percent+label',
        hovertemplate='<b>%{label}</b><br>Počet: %{value}<br>Podiel: %{percent}<extra></extra>'
    )
    
    fig.update_layout(height=400)
    st.plotly_chart(fig, use_container_width=True)

def render_match_strategy_chart(results: Dict[str, List]):
    """
    Vykresli graf stratégií zhody.
    """
    if 'MatchStrategy' not in results:
        st.info("Údaje o stratégiách zhody nie sú dostupné")
        return
    
    strategies = [s for s in results['MatchStrategy'] if s]
    if not strategies:
        st.info("Žiadne údaje o stratégiách zhody")
        return
    
    strategy_counts = pd.Series(strategies).value_counts()
    
    fig = px.bar(
        x=strategy_counts.index,
        y=strategy_counts.values,
        title='Použité stratégie zhody',
        labels={'x': 'Stratégia', 'y': 'Počet použití'},
        color=strategy_counts.values,
        color_continuous_scale='Viridis'
    )
    
    fig.update_traces(
        hovertemplate='<b>%{x}</b><br>Počet: %{y}<extra></extra>'
    )
    
    fig.update_layout(height=400, xaxis_title="Stratégia zhody", yaxis_title="Počet použití")
    st.plotly_chart(fig, use_container_width=True)

def render_advanced_analytics(results: Dict[str, List], df: pd.DataFrame):
    """
    Pokročilé analýzy výsledkov.
    """
    st.subheader("🔍 Pokročilé analýzy")
    
    tabs = st.tabs(["📊 Kvalita zhôd", "🔧 Technické detaily", "📈 Trendy", "❌ Analýza chýb"])
    
    with tabs[0]:
        render_match_quality_analysis(results)
    
    with tabs[1]:
        render_technical_details(results)
    
    with tabs[2]:
        render_trends_analysis(results, df)
    
    with tabs[3]:
        render_error_analysis(results)

def render_match_quality_analysis(results: Dict[str, List]):
    """
    Analýza kvality zhôd.
    """
    if 'MatchStrategy' not in results or 'ICO' not in results:
        st.info("Nedostatok dát pre analýzu kvality")
        return
    
    # Analýza úspešnosti podľa stratégie
    strategy_success = {}
    for i, strategy in enumerate(results['MatchStrategy']):
        if strategy:
            ico = results['ICO'][i] if i < len(results['ICO']) else None
            if strategy not in strategy_success:
                strategy_success[strategy] = {'total': 0, 'successful': 0}
            strategy_success[strategy]['total'] += 1
            if ico:
                strategy_success[strategy]['successful'] += 1
    
    if strategy_success:
        strategy_df = pd.DataFrame([
            {
                'Stratégia': strategy,
                'Celkovo': data['total'],
                'Úspešné': data['successful'],
                'Úspešnosť (%)': (data['successful'] / data['total']) * 100
            }
            for strategy, data in strategy_success.items()
        ])
        
        fig = px.bar(
            strategy_df,
            x='Stratégia',
            y='Úspešnosť (%)',
            color='Úspešnosť (%)',
            title='Úspešnosť jednotlivých stratégií zhody',
            color_continuous_scale='RdYlGn',
            text='Úspešnosť (%)'
        )
        
        fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
        
        # Tabuľka s detailami
        st.dataframe(strategy_df, use_container_width=True)

def render_technical_details(results: Dict[str, List]):
    """
    Technické detaily o spracovaní.
    """
    # Analýza typov identifikátorov
    if 'IdentifierType' in results:
        id_types = [t for t in results['IdentifierType'] if t]
        if id_types:
            type_counts = pd.Series(id_types).value_counts()
            
            col1, col2 = st.columns(2)
            
            with col1:
                fig = px.pie(
                    values=type_counts.values,
                    names=type_counts.index,
                    title='Typy identifikátorov'
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with col2:
                st.markdown("**Detaily typov identifikátorov:**")
                for id_type, count in type_counts.items():
                    percentage = (count / len(id_types)) * 100
                    st.write(f"• **{id_type}**: {count} ({percentage:.1f}%)")
    
    # Analýza variantov dotazov
    if 'UsedQueryVariant' in results:
        variants = [v for v in results['UsedQueryVariant'] if v]
        if variants:
            # Top 10 najpoužívanejších variantov
            variant_counts = pd.Series(variants).value_counts().head(10)
            
            st.markdown("**Top 10 najpoužívanejších variantov dotazov:**")
            variant_df = pd.DataFrame({
                'Variant': variant_counts.index,
                'Počet použití': variant_counts.values
            })
            st.dataframe(variant_df, use_container_width=True)

def render_trends_analysis(results: Dict[str, List], df: pd.DataFrame):
    """
    Analýza trendov v dátach.
    """
    # Analýza podľa dĺžky názvu firmy
    if 'CleanName' in results and 'ICO' in results:
        name_lengths = [len(name) if name else 0 for name in results['CleanName']]
        success_by_length = []
        
        # Zoskupenie podľa dĺžky
        length_bins = range(0, max(name_lengths) + 10, 10)
        
        for i in range(len(length_bins) - 1):
            min_len = length_bins[i]
            max_len = length_bins[i + 1]
            
            in_bin = [j for j, length in enumerate(name_lengths) if min_len <= length < max_len]
            if in_bin:
                successful_in_bin = sum(1 for j in in_bin if results['ICO'][j])
                success_rate = (successful_in_bin / len(in_bin)) * 100
                
                success_by_length.append({
                    'Dĺžka názvu': f"{min_len}-{max_len}",
                    'Počet': len(in_bin),
                    'Úspešné': successful_in_bin,
                    'Úspešnosť (%)': success_rate
                })
        
        if success_by_length:
            trends_df = pd.DataFrame(success_by_length)
            
            fig = px.line(
                trends_df,
                x='Dĺžka názvu',
                y='Úspešnosť (%)',
                title='Úspešnosť podľa dĺžky názvu firmy',
                markers=True
            )
            
            fig.update_layout(height=400)
            st.plotly_chart(fig, use_container_width=True)
            
            st.dataframe(trends_df, use_container_width=True)

def render_error_analysis(results: Dict[str, List]):
    """
    Detailná analýza chýb a problémov.
    """
    if 'Notes' not in results:
        st.info("Údaje o chybách nie sú dostupné")
        return
    
    notes = [note for note in results['Notes'] if note]
    
    if not notes:
        st.success("✅ Žiadne chyby neboli zaznamenané")
        return
    
    # Kategorizácia chýb
    error_categories = {
        'API chyby': [],
        'Nenájdené firmy': [],
        'Neplatné IČO': [],
        'Technické problémy': [],
        'Ostatné': []
    }
    
    for note in notes:
        if 'Exception' in note:
            error_categories['Technické problémy'].append(note)
        elif 'Not found' in note:
            error_categories['Nenájdené firmy'].append(note)
        elif 'without valid ICO' in note:
            error_categories['Neplatné IČO'].append(note)
        elif 'API' in note or 'timeout' in note.lower():
            error_categories['API chyby'].append(note)
        else:
            error_categories['Ostatné'].append(note)
    
    # Graf kategórií chýb
    category_counts = {k: len(v) for k, v in error_categories.items() if v}
    
    if category_counts:
        fig = px.bar(
            x=list(category_counts.keys()),
            y=list(category_counts.values()),
            title='Rozdelenie typov chýb',
            color=list(category_counts.values()),
            color_continuous_scale='Reds'
        )
        
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
        
        # Detailné zobrazenie chýb
        for category, errors in error_categories.items():
            if errors:
                with st.expander(f"🔍 {category} ({len(errors)})", expanded=False):
                    for i, error in enumerate(errors[:10], 1):  # Limit na 10
                        st.write(f"{i}. {error}")
                    if len(errors) > 10:
                        st.info(f"... a ďalších {len(errors) - 10} chýb")

def render_export_section(results: Dict[str, List], df: pd.DataFrame):
    """
    Sekcia pre export výsledkov.
    """
    st.subheader("💾 Export výsledkov")
    
    col1, col2, col3 = st.columns(3)
    
    # Príprava output DataFrame
    output_df = df.copy()
    for key, values in results.items():
        # Zabezpečenie správnej dĺžky
        while len(values) < len(output_df):
            values.append(None)
        output_df[key] = values[:len(output_df)]
    
    with col1:
        # Excel export s formátovaním
        excel_buffer = create_formatted_excel(output_df)
        st.download_button(
            label="📊 Excel s formátovaním",
            data=excel_buffer,
            file_name=f"ico_results_formatted_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Excel súbor s farebným kódovaním a formátovaním"
        )
    
    with col2:
        # CSV export
        csv_data = output_df.to_csv(index=False, encoding='utf-8')
        st.download_button(
            label="📋 CSV súbor",
            data=csv_data,
            file_name=f"ico_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            help="CSV súbor pre ďalšie spracovanie"
        )
    
    with col3:
        # JSON export pre pokročilých používateľov
        json_data = output_df.to_json(orient='records', indent=2, force_ascii=False)
        st.download_button(
            label="📄 JSON dáta",
            data=json_data,
            file_name=f"ico_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            help="JSON formát pre programové použitie"
        )

def create_formatted_excel(df: pd.DataFrame):
    """
    Vytvorí Excel súbor s pokročilým formátovaním.
    """
    from io import BytesIO
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='ICO_Výsledky', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['ICO_Výsledky']
        
        # Formátovanie hlavičky
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Formátovanie ICO stĺpca
        if 'ICO' in df.columns:
            ico_col_idx = df.columns.get_loc('ICO') + 1
            success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=ico_col_idx)
                if cell.value:
                    cell.fill = success_fill
                else:
                    cell.fill = error_fill
        
        # Automatické prispôsobenie šírky stĺpcov
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output

def render_comparison_dashboard(results_list: List[Dict], names: List[str]):
    """
    Dashboard pre porovnanie viacerých spracovaní.
    """
    if len(results_list) < 2:
        st.info("Potrebujete aspoň 2 spracovávania na porovnanie")
        return
    
    st.subheader("🆚 Porovnanie spracovaní")
    
    comparison_data = []
    for i, (results, name) in enumerate(zip(results_list, names)):
        total = len(results.get('ICO', []))
        successful = sum(1 for ico in results.get('ICO', []) if ico)
        success_rate = (successful / total) * 100 if total > 0 else 0
        
        comparison_data.append({
            'Spracovanie': name,
            'Celkovo': total,
            'Úspešné': successful,
            'Úspešnosť (%)': success_rate
        })
    
    comparison_df = pd.DataFrame(comparison_data)
    
    # Graf porovnania
    fig = px.bar(
        comparison_df,
        x='Spracovanie',
        y='Úspešnosť (%)',
        color='Úspešnosť (%)',
        title='Porovnanie úspešnosti spracovaní',
        text='Úspešnosť (%)'
    )
    
    fig.update_traces(texttemplate='%{text:.1f}%', textposition='outside')
    st.plotly_chart(fig, use_container_width=True)
    
    # Tabuľka porovnania
    st.dataframe(comparison_df, use_container_width=True)